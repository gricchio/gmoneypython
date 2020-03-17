'''
Created on Mar 5, 2020

@author: Gino.Ricchio
'''
#Imports
import os 
import openpyxl
#hello
#Variables
project_folder = r"C:\Users\gino.ricchio\Desktop\Python Projects\Delete Tabs"
inactive = ['ACCOUNTING',
            'ADMIN', 
            'BLANK', 
            'DAIRY', 
            'DC', 
            'DEV', 
            'EQUIP', 
            'FEED', 
            'FW', 
            'G & A', 
            'GENERAL SW', 
            'M & S', 
            'MARKETING', 
            'MGMT', 
            'MORPHEUS', 
            'NA', 
            'OFFICE', 
            'PARLOR', 
            'PCC', 
            'PLATFORM', 
            'PW', 
            'SPEC PROJ', 
            'TAGS', 
            'TLINK', 
            'VAS LINKS', 
            'VVILEASED', 
            'WEIGHRITE',
            'WIRELESS',
            ]

#Code


os.chdir(project_folder)


files = []
for name in os.listdir(project_folder):
    if os.path.isfile(os.path.join(project_folder, name)):
        files.append(name)

workbook = openpyxl.load_workbook(files[0])


for sheet in workbook.sheetnames:
    ws = workbook[sheet]
    if sheet in inactive:
        ws.sheet_state = 'hidden'
        print(sheet, '---Hidden')
print(workbook.sheetnames)

workbook.save(str(files[0]))

